import json
import os
import subprocess
import sys
import time
import traceback
import zipfile
from pathlib import Path
from typing import Any, Literal

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from pydantic import Field
from pydantic.fields import FieldInfo

from ..base import ActionArguments, ActionCollection, ActionResponse
from ..documents.models import DocumentMetadata


class XLSXExtractionCollection(ActionCollection):
    """MCP service for Excel document content extraction using xlrd and pandas.

    Supports extraction from XLSX and XLS files.
    Provides LLM-friendly text output with structured metadata and media file handling.
    Extracts worksheets, formulas, charts, and embedded images.
    Includes screenshot functionality for visual representation of Excel data.
    """
    tool_name = "xlsx"

    def __init__(self, arguments: ActionArguments) -> None:
        super().__init__(arguments)
        self._media_output_dir = self.workspace / "extracted_media"
        self._media_output_dir.mkdir(exist_ok=True)

        # Create screenshots directory
        self._screenshots_dir = self.workspace / "excel_screenshots"
        self._screenshots_dir.mkdir(exist_ok=True)

        self.supported_extensions: set = {
            ".xlsx",
            ".xls",
        }

        print("Excel Extraction Service initialized")
        print(f"Media output directory: {self._media_output_dir}")
        print(f"Screenshots directory: {self._screenshots_dir}")

    def _create_excel_screenshot(self, file_path: Path, sheet_name: str = None) -> str:
        """Create a JPEG screenshot of the valid Excel area using pyautogui.

        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet to screenshot (None for first sheet)

        Returns:
            Path to the generated JPEG screenshot
        """
        try:
            import pyautogui

            # Generate unique filename
            timestamp = int(time.time())
            screenshot_filename = f"{file_path.stem}_{sheet_name or 'sheet'}_{timestamp}.jpg"
            screenshot_path = self._screenshots_dir / screenshot_filename

            # Open Excel file with default application
            if sys.platform == "darwin":  # macOS
                subprocess.run(["open", str(file_path)], check=True)
            elif sys.platform == "win32":  # Windows
                subprocess.run(["start", str(file_path)], shell=True, check=True)
            else:  # Linux
                subprocess.run(["xdg-open", str(file_path)], check=True)

            # Wait for Excel to open
            time.sleep(3)

            # Take screenshot of the entire screen
            screenshot = pyautogui.screenshot()

            # Convert RGBA to RGB before saving as JPEG
            if screenshot.mode == "RGBA":
                screenshot = screenshot.convert("RGB")

            screenshot.save(screenshot_path, "JPEG", quality=95)

            print(f"Created Excel screenshot: {screenshot_filename}")
            return str(screenshot_path)

        except Exception as e:
            self.logger.error(f"Failed to create Excel screenshot with pyautogui: {str(e)}")
            raise

    def _extract_embedded_media_xlsx(self, file_path: Path) -> list[dict[str, str]]:
        """Extract embedded media from XLSX files.

        Args:
            file_path: Path to the XLSX file

        Returns:
            List of dictionaries containing media information
        """
        saved_media = []

        try:
            # Load workbook to extract images
            workbook = load_workbook(file_path, data_only=False)

            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Extract images from worksheet
                if hasattr(worksheet, "_images"):
                    for idx, image in enumerate(worksheet._images):
                        try:
                            # Generate unique filename
                            image_filename = f"{file_path.stem}_{sheet_name}_img_{idx}.png"
                            image_path = self._media_output_dir / image_filename

                            # Save image
                            if hasattr(image, "ref"):
                                # Extract image data
                                img_data = image._data()
                                if img_data:
                                    with open(image_path, "wb") as f:
                                        f.write(img_data)

                                    saved_media.append(
                                        {
                                            "type": "image",
                                            "path": str(image_path),
                                            "sheet": sheet_name,
                                            "filename": image_filename,
                                        }
                                    )

                                    print(f"Saved image: {image_filename}")
                        except Exception as e:
                            self.logger.warning(f"Failed to extract image {idx} from sheet {sheet_name}: {str(e)}")

            # Also try to extract from ZIP structure for additional media
            with zipfile.ZipFile(file_path, "r") as zip_file:
                media_files = [f for f in zip_file.namelist() if f.startswith("xl/media/")]

                for media_file in media_files:
                    try:
                        media_data = zip_file.read(media_file)
                        media_filename = f"{file_path.stem}_{Path(media_file).name}"
                        media_path = self._media_output_dir / media_filename

                        with open(media_path, "wb") as f:
                            f.write(media_data)

                        # Determine media type based on extension
                        media_ext = Path(media_file).suffix.lower()
                        if media_ext in [".png", ".jpg", ".jpeg", ".gif", ".bmp"]:
                            media_type = "image"
                        elif media_ext in [".mp3", ".wav", ".m4a"]:
                            media_type = "audio"
                        elif media_ext in [".mp4", ".avi", ".mov"]:
                            media_type = "video"
                        else:
                            media_type = "other"

                        saved_media.append(
                            {
                                "type": media_type,
                                "path": str(media_path),
                                "filename": media_filename,
                                "original_path": media_file,
                            }
                        )

                        print(f"Saved media: {media_filename}")
                    except Exception as e:
                        self.logger.warning(f"Failed to extract media {media_file}: {str(e)}")

        except Exception as e:
            self.logger.warning(f"Failed to extract media from XLSX: {str(e)}")

        return saved_media

    def _extract_excel_content(self, file_path: Path, sheet_names: list[str] | None = None) -> dict[str, Any]:
        """Extract content from Excel files using pandas and xlrd.

        Args:
            file_path: Path to the Excel file
            sheet_names: Specific sheets to process (None for all sheets)

        Returns:
            Dictionary containing extracted content and metadata
        """
        start_time = time.time()
        timeout_seconds = 60  # 60 second timeout
        
        try:
            sys.stderr.write(f"Starting Excel processing for {file_path}\n")
            
            # Add a simple timeout check during processing
            def check_timeout():
                if time.time() - start_time > timeout_seconds:
                    sys.stderr.write(f"Excel processing timed out after {timeout_seconds} seconds\n")
                    raise TimeoutError(f"Excel processing timed out after {timeout_seconds} seconds")
            
            # Determine file type and read accordingly
            if file_path.suffix.lower() == ".xlsx":
                # Use openpyxl engine for XLSX files
                sys.stderr.write("Using openpyxl engine for XLSX file\n")
                excel_file = pd.ExcelFile(file_path, engine="openpyxl")
            else:
                # Use xlrd engine for XLS files
                sys.stderr.write("Using xlrd engine for XLS file\n")
                excel_file = pd.ExcelFile(file_path, engine="xlrd")

            check_timeout()  # Check timeout after loading file

            # Get all sheet names if not specified
            if sheet_names is None:
                sheet_names = excel_file.sheet_names
                sys.stderr.write(f"Found sheets: {', '.join(sheet_names)}\n")
            else:
                sys.stderr.write(f"Using specified sheets: {', '.join(sheet_names)}\n")

            sheets_data = {}
            total_rows = 0
            total_cols = 0

            # Extract data from each sheet
            for sheet_name in sheet_names:
                check_timeout()  # Check timeout before processing each sheet
                
                if sheet_name in excel_file.sheet_names:
                    try:
                        sys.stderr.write(f"Reading sheet: {sheet_name}\n")
                        # Read sheet data
                        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

                        # Remove completely empty rows and columns
                        df = df.dropna(how="all").dropna(axis=1, how="all")

                        if not df.empty:
                            sheets_data[sheet_name] = {
                                "data": df,
                                "shape": df.shape,
                                "columns": df.columns.tolist(),
                                "non_empty_cells": df.count().sum(),
                            }

                            total_rows += df.shape[0]
                            total_cols = max(total_cols, df.shape[1])
                            sys.stderr.write(f"Sheet {sheet_name}: {df.shape[0]} rows, {df.shape[1]} columns\n")
                        else:
                            sheets_data[sheet_name] = {"data": df, "shape": (0, 0), "columns": [], "non_empty_cells": 0}
                            sys.stderr.write(f"Sheet {sheet_name} is empty\n")

                    except Exception as e:
                        sys.stderr.write(f"Error reading sheet '{sheet_name}': {str(e)}\n")
                        sheets_data[sheet_name] = {
                            "error": str(e),
                            "shape": (0, 0),
                            "columns": [],
                            "non_empty_cells": 0,
                        }

            processing_time = time.time() - start_time
            sys.stderr.write(f"Excel processing completed in {processing_time:.2f} seconds\n")

            return {
                "sheets_data": sheets_data,
                "sheet_names": list(sheets_data.keys()),
                "total_sheets": len(sheets_data),
                "total_rows": total_rows,
                "total_columns": total_cols,
                "processing_time": processing_time,
                "file_engine": "openpyxl" if file_path.suffix.lower() == ".xlsx" else "xlrd",
            }

        except TimeoutError as e:
            sys.stderr.write(f"Excel processing timed out: {str(e)}\n")
            return {
                "sheets_data": {},
                "sheet_names": [],
                "total_sheets": 0,
                "total_rows": 0,
                "total_columns": 0,
                "processing_time": time.time() - start_time,
                "error": str(e),
                "file_engine": "openpyxl" if file_path.suffix.lower() == ".xlsx" else "xlrd",
            }
        except Exception as e:
            sys.stderr.write(f"Excel processing failed: {str(e)}\n")
            if hasattr(e, "__traceback__"):
                sys.stderr.write(f"{traceback.format_exc()}\n")
            raise

    def _format_content_for_llm(
        self, extraction_result: dict[str, Any], output_format: str, include_empty_cells: bool = False
    ) -> str:
        """Format extracted Excel content to be LLM-friendly.

        Args:
            extraction_result: Result from _extract_excel_content
            output_format: Desired output format
            include_empty_cells: Whether to include empty cells in output

        Returns:
            Formatted content string
        """
        sheets_data = extraction_result["sheets_data"]
        max_rows_per_sheet = 100  # Limit rows to prevent excessive output

        if output_format.lower() == "markdown":
            content_parts = []
            content_parts.append("# Excel Document Content\n")
            content_parts.append(f"**Total Sheets:** {extraction_result['total_sheets']}\n")
            content_parts.append(f"**Processing Engine:** {extraction_result['file_engine']}\n\n")

            for sheet_name, sheet_info in sheets_data.items():
                content_parts.append(f"## Sheet: {sheet_name}\n")
                
                # Check for error in sheet
                if "error" in sheet_info:
                    content_parts.append(f"**Error:** {sheet_info['error']}\n\n")
                    continue
                    
                # Get sheet dimensions
                rows, cols = sheet_info["shape"]
                content_parts.append(f"**Dimensions:** {rows} rows × {cols} columns\n")
                content_parts.append(f"**Non-empty cells:** {sheet_info['non_empty_cells']}\n\n")

                # Check if sheet has data
                if rows == 0 or cols == 0:
                    content_parts.append("*Empty sheet*\n\n")
                    continue

                # Get data frame with rows limited
                df = sheet_info["data"]
                if len(df) > max_rows_per_sheet:
                    content_parts.append(f"*Showing first {max_rows_per_sheet} rows of {len(df)} total rows*\n\n")
                    df = df.iloc[:max_rows_per_sheet]

                # Format as markdown table
                try:
                    # Reset index to get proper row numbering
                    df_display = df.reset_index(drop=True)
                    
                    # Drop completely empty rows/columns if requested
                    if not include_empty_cells:
                        df_display = df_display.dropna(how="all").dropna(axis=1, how="all")
                    
                    # Generate markdown table
                    table_str = df_display.to_markdown(index=True)
                    content_parts.append(f"{table_str}\n\n")
                except Exception as e:
                    sys.stderr.write(f"Error formatting sheet {sheet_name} as markdown: {str(e)}\n")
                    # Fallback to simple text representation
                    content_parts.append("*Error formatting table - showing sample values:*\n\n")
                    content_parts.append("```\n")
                    content_parts.append(str(df.iloc[:5, :5]) + "\n")
                    content_parts.append("...\n")
                    content_parts.append("```\n\n")

            return "".join(content_parts)
            
        elif output_format.lower() == "json":
            # For JSON output, prepare a simplified representation
            sheets_json = {}
            for sheet_name, sheet_info in sheets_data.items():
                if "error" in sheet_info:
                    sheets_json[sheet_name] = {"error": sheet_info["error"]}
                    continue
                    
                df = sheet_info["data"]
                if len(df) > max_rows_per_sheet:
                    df = df.iloc[:max_rows_per_sheet]
                
                try:
                    # Convert to JSON-safe format
                    sheets_json[sheet_name] = {
                        "dimensions": sheet_info["shape"],
                        "non_empty_cells": sheet_info["non_empty_cells"],
                        "data": json.loads(df.to_json(orient="records")),
                        "truncated": len(df) < sheet_info["shape"][0]
                    }
                except Exception as e:
                    sheets_json[sheet_name] = {
                        "dimensions": sheet_info["shape"],
                        "non_empty_cells": sheet_info["non_empty_cells"],
                        "error": f"Error converting to JSON: {str(e)}",
                    }
            
            output_json = {
                "document_info": {
                    "total_sheets": extraction_result["total_sheets"],
                    "total_rows": extraction_result["total_rows"],
                    "total_columns": extraction_result["total_columns"],
                    "processing_engine": extraction_result["file_engine"],
                    "processing_time": extraction_result["processing_time"],
                },
                "sheets": sheets_json
            }
            
            return json.dumps(output_json, indent=2)
            
        elif output_format.lower() == "html":
            html_parts = []
            html_parts.append("<html><body>")
            html_parts.append("<h1>Excel Document Content</h1>")
            html_parts.append(f"<p><strong>Total Sheets:</strong> {extraction_result['total_sheets']}</p>")
            
            for sheet_name, sheet_info in sheets_data.items():
                html_parts.append(f"<h2>Sheet: {sheet_name}</h2>")
                
                if "error" in sheet_info:
                    html_parts.append(f"<p class='error'>Error: {sheet_info['error']}</p>")
                    continue
                    
                rows, cols = sheet_info["shape"]
                html_parts.append(f"<p>Dimensions: {rows} rows × {cols} columns</p>")
                
                if rows == 0 or cols == 0:
                    html_parts.append("<p><em>Empty sheet</em></p>")
                    continue
                    
                df = sheet_info["data"]
                if len(df) > max_rows_per_sheet:
                    html_parts.append(f"<p><em>Showing first {max_rows_per_sheet} rows of {len(df)} total rows</em></p>")
                    df = df.iloc[:max_rows_per_sheet]
                    
                try:
                    # Reset index for proper display
                    df_display = df.reset_index(drop=True)
                    if not include_empty_cells:
                        df_display = df_display.dropna(how="all").dropna(axis=1, how="all")
                        
                    # Convert to HTML table
                    table_html = df_display.to_html(index=True)
                    html_parts.append(table_html)
                except Exception as e:
                    html_parts.append(f"<p>Error formatting table: {str(e)}</p>")
                    html_parts.append("<pre>")
                    html_parts.append(str(df.iloc[:5, :5]))
                    html_parts.append("...</pre>")
            
            html_parts.append("</body></html>")
            return "".join(html_parts)
            
        else:  # Plain text
            text_parts = []
            text_parts.append("Excel Document Content\n")
            text_parts.append("=" * 40 + "\n\n")
            text_parts.append(f"Total Sheets: {extraction_result['total_sheets']}\n")
            text_parts.append(f"Processing Engine: {extraction_result['file_engine']}\n\n")
            
            for sheet_name, sheet_info in sheets_data.items():
                text_parts.append(f"Sheet: {sheet_name}\n")
                text_parts.append("-" * 40 + "\n")
                
                if "error" in sheet_info:
                    text_parts.append(f"Error: {sheet_info['error']}\n\n")
                    continue
                    
                rows, cols = sheet_info["shape"]
                text_parts.append(f"Dimensions: {rows} rows × {cols} columns\n")
                text_parts.append(f"Non-empty cells: {sheet_info['non_empty_cells']}\n\n")
                
                if rows == 0 or cols == 0:
                    text_parts.append("Empty sheet\n\n")
                    continue
                    
                df = sheet_info["data"]
                if len(df) > max_rows_per_sheet:
                    text_parts.append(f"Showing first {max_rows_per_sheet} rows of {len(df)} total rows\n\n")
                    df = df.iloc[:max_rows_per_sheet]
                
                # Simple text representation
                try:
                    df_display = df.reset_index(drop=True)
                    if not include_empty_cells:
                        df_display = df_display.dropna(how="all").dropna(axis=1, how="all")
                        
                    text_parts.append(str(df_display) + "\n\n")
                except Exception as e:
                    text_parts.append(f"Error formatting table: {str(e)}\n")
                    text_parts.append(str(df.iloc[:5, :5]) + "\n...\n\n")
                    
            return "".join(text_parts)

    def _extract_cell_colors(self, file_path: Path, sheet_name: str = None) -> dict:
        """Extract cell background colors from Excel file using openpyxl.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet name (None for first sheet)
            
        Returns:
            Dictionary mapping cell coordinates to their background colors
        """
        try:
            # Only works with xlsx files using openpyxl
            if file_path.suffix.lower() != ".xlsx":
                return {"error": "Color extraction only supported for .xlsx files"}
                
            # Load workbook with openpyxl directly
            from openpyxl import load_workbook
            
            sys.stderr.write(f"Extracting cell colors from {file_path}\n")
            wb = load_workbook(file_path)
            
            # Select sheet
            if sheet_name and sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                sheet = wb.active
                
            cell_colors = {}
            
            # Iterate through cells
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    
                    # Get cell value and coordinates
                    cell_value = cell.value
                    cell_coord = f"{row},{col}"
                    
                    # Get cell background color if available
                    bg_color = None
                    if cell.fill and hasattr(cell.fill, 'start_color') and hasattr(cell.fill.start_color, 'rgb'):
                        bg_color = cell.fill.start_color.rgb
                        # Convert AARRGGBB format to standard hex
                        if bg_color and len(bg_color) == 8:
                            bg_color = bg_color[2:]  # Remove alpha channel
                    
                    cell_colors[cell_coord] = {
                        "value": cell_value,
                        "color": bg_color
                    }
            
            return cell_colors
            
        except Exception as e:
            sys.stderr.write(f"Error extracting cell colors: {str(e)}\n{traceback.format_exc()}\n")
            return {"error": f"Failed to extract colors: {str(e)}"}
            
    def mcp_extract_cell_colors(
        self,
        file_path: str = Field(description="Path to the Excel document file to extract colors from"),
        sheet_name: str | None = Field(
            default=None, description="Specific sheet name to extract colors from (None for active sheet)"
        ),
    ) -> ActionResponse:
        """Extract cell background colors from Excel spreadsheets.
        
        This tool extracts background colors from Excel cells, which is useful for:
        - Color-coded maps or grids
        - Analyzing visual patterns in spreadsheets
        - Path-finding problems that rely on cell colors
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet to process
            
        Returns:
            ActionResponse with cell coordinates mapped to colors
        """
        try:
            # Handle FieldInfo objects
            if isinstance(file_path, FieldInfo):
                file_path = file_path.default
            if isinstance(sheet_name, FieldInfo):
                sheet_name = sheet_name.default
                
            # Validate input file
            file_path: Path = self._validate_file_path(file_path)
            sys.stderr.write(f"Processing Excel document for color extraction: {file_path.name}\n")
            
            # Extract colors
            cell_colors = self._extract_cell_colors(file_path, sheet_name)
            
            # Check if error occurred
            if "error" in cell_colors:
                return ActionResponse(
                    success=False,
                    message=f"Color extraction failed: {cell_colors['error']}",
                    metadata={"error_type": "extraction_error"},
                )
                
            # Count cells with non-default colors
            colored_cells = sum(1 for info in cell_colors.values() if info.get("color") and info["color"] != "FFFFFF")
            
            # Prepare metadata
            file_stats = file_path.stat()
            metadata = {
                "file_name": file_path.name,
                "file_size": file_stats.st_size,
                "file_type": file_path.suffix.lower(),
                "total_cells": len(cell_colors),
                "colored_cells": colored_cells,
                "sheet_name": sheet_name or "default"
            }
            
            # Create message content
            message = f"Successfully extracted colors from {len(cell_colors)} cells in {file_path.name}."
            if colored_cells > 0:
                message += f" Found {colored_cells} cells with non-default background colors."
            
            # Convert cell colors to a more readable format for the response
            formatted_cells = {}
            for coord, info in cell_colors.items():
                formatted_cells[coord] = {
                    "value": str(info["value"]) if info["value"] is not None else "Empty",
                    "color": info["color"] if info["color"] else "default"
                }
            
            metadata["cell_colors"] = formatted_cells
            
            return ActionResponse(
                success=True,
                message=message,
                metadata=metadata
            )
            
        except FileNotFoundError as e:
            sys.stderr.write(f"File not found: {str(e)}\n")
            return ActionResponse(
                success=False, 
                message=f"File not found: {str(e)}", 
                metadata={"error_type": "file_not_found"}
            )
        except Exception as e:
            sys.stderr.write(f"Color extraction failed: {str(e)}\n{traceback.format_exc()}\n")
            return ActionResponse(
                success=False,
                message=f"Color extraction failed: {str(e)}",
                metadata={"error_type": "extraction_error"},
            )

    def _validate_file_path(self, file_path: str) -> Path:
        """Validate and resolve file path.

        Args:
            file_path: Path to the Excel document file

        Returns:
            Resolved Path object

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file type is not supported
        """
        # Convert to Path object
        path = Path(file_path)
        
        # Handle relative paths
        if not path.is_absolute():
            # Try current directory first
            if not path.exists():
                # Try workspace directory
                workspace_path = Path(self.workspace) / path
                if workspace_path.exists():
                    path = workspace_path
                else:
                    # Try expanded path
                    expanded_path = Path(os.path.expanduser(file_path))
                    if expanded_path.exists():
                        path = expanded_path
        
        sys.stderr.write(f"Validating file path: {path}\n")
        
        # Check if file exists
        if not path.exists():
            error_msg = f"File not found: {path}"
            sys.stderr.write(f"ERROR: {error_msg}\n")
            raise FileNotFoundError(error_msg)
        
        # Check file extension
        if path.suffix.lower() not in self.supported_extensions:
            error_msg = f"Unsupported file type: {path.suffix}. Supported types: {', '.join(self.supported_extensions)}"
            sys.stderr.write(f"ERROR: {error_msg}\n")
            raise ValueError(error_msg)
        
        sys.stderr.write(f"File validated successfully: {path.absolute()}\n")
        return path

    def mcp_extract_excel_content(
        self,
        file_path: str = Field(description="Path to the Excel document file to extract content from"),
        output_format: Literal["markdown", "json", "html", "text"] = Field(
            default="markdown", description="Output format: 'markdown', 'json', 'html', or 'text'"
        ),
        extract_images: bool = Field(default=False, description="Whether to extract and save images from the document"),
        create_screenshot: bool = Field(
            default=False, description="Whether to create a JPEG screenshot of the Excel data"
        ),
        sheet_names: str | None = Field(
            default=None, description="Comma-separated list of specific sheet names to process (None for all sheets)"
        ),
        include_empty_cells: bool = Field(default=False, description="Whether to include empty cells in the output"),
        screenshot_max_rows: int = Field(default=50, description="Maximum rows to include in screenshot"),
        screenshot_max_cols: int = Field(default=20, description="Maximum columns to include in screenshot"),
    ) -> ActionResponse:
        """Extract content from Excel documents using pandas and xlrd.

        This tool provides comprehensive Excel document content extraction with support for:
        - XLSX and XLS files
        - Multiple worksheets
        - Text and numeric data extraction
        - Image and media extraction (XLSX only)
        - JPEG screenshot generation of Excel data
        - Metadata collection
        - LLM-optimized output formatting
        """
        try:
            # Handle FieldInfo objects
            if isinstance(file_path, FieldInfo):
                file_path = file_path.default
            if isinstance(output_format, FieldInfo):
                output_format = output_format.default
            if isinstance(extract_images, FieldInfo):
                extract_images = extract_images.default
            if isinstance(create_screenshot, FieldInfo):
                create_screenshot = create_screenshot.default
            if isinstance(sheet_names, FieldInfo):
                sheet_names = sheet_names.default
            if isinstance(include_empty_cells, FieldInfo):
                include_empty_cells = include_empty_cells.default
            if isinstance(screenshot_max_rows, FieldInfo):
                screenshot_max_rows = screenshot_max_rows.default
            if isinstance(screenshot_max_cols, FieldInfo):
                screenshot_max_cols = screenshot_max_cols.default

            # Force create_screenshot to False for safety in server environments
            create_screenshot = False
            
            # Validate input file
            file_path: Path = self._validate_file_path(file_path)
            sys.stderr.write(f"Processing Excel document: {file_path.name}\n")

            # Parse sheet names if provided
            target_sheets = None
            if sheet_names:
                target_sheets = [name.strip() for name in sheet_names.split(",")]
                sys.stderr.write(f"Target sheets: {target_sheets}\n")

            # Extract content from Excel file
            extraction_result = self._extract_excel_content(file_path, target_sheets)

            # Extract embedded media if requested (XLSX only) - but handle errors gracefully
            saved_media = []
            if extract_images and file_path.suffix.lower() == ".xlsx":
                try:
                    saved_media = self._extract_embedded_media_xlsx(file_path)
                except Exception as e:
                    sys.stderr.write(f"Failed to extract images: {str(e)}\n")
                    saved_media = []
            elif extract_images and file_path.suffix.lower() == ".xls":
                sys.stderr.write("Image extraction not supported for XLS files\n")

            # Format content for LLM consumption
            formatted_content = self._format_content_for_llm(extraction_result, output_format, include_empty_cells)

            # Prepare metadata
            file_stats = file_path.stat()

            # Create Excel-specific metadata
            excel_metadata = {
                "sheet_count": extraction_result["total_sheets"],
                "sheet_names": extraction_result["sheet_names"],
                "total_rows": extraction_result["total_rows"],
                "total_columns": extraction_result["total_columns"],
                "processing_engine": extraction_result["file_engine"],
                "extracted_images": [media["path"] for media in saved_media if media["type"] == "image"],
                "extracted_media": saved_media,
            }

            document_metadata = DocumentMetadata(
                file_name=file_path.name,
                file_size=file_stats.st_size,
                file_type=file_path.suffix.lower(),
                absolute_path=str(file_path.absolute()),
                page_count=len(extraction_result["sheet_names"]),  # Use sheet count as page count
                processing_time=extraction_result["processing_time"],
                extracted_images=[media["path"] for media in saved_media if media["type"] == "image"],
                extracted_media=saved_media,
                output_format=output_format,
                llm_enhanced=False,
                ocr_applied=False,
            )

            # Combine standard and Excel-specific metadata
            combined_metadata = document_metadata.model_dump()
            combined_metadata.update(excel_metadata)

            sys.stderr.write(
                f"Successfully extracted content from {file_path.name} "
                f"({extraction_result['total_sheets']} sheets, "
                f"({extraction_result['total_rows']} rows, "
                f"{len(saved_media)} media files)\n"
            )

            return ActionResponse(success=True, message=formatted_content, metadata=combined_metadata)

        except FileNotFoundError as e:
            sys.stderr.write(f"File not found: {str(e)}\n")
            return ActionResponse(
                success=False, message=f"File not found: {str(e)}", metadata={"error_type": "file_not_found"}
            )
        except ValueError as e:
            sys.stderr.write(f"Invalid input: {str(e)}\n")
            return ActionResponse(
                success=False,
                message=f"Invalid input: {str(e)}",
                metadata={"error_type": "invalid_input"},
            )
        except Exception as e:
            sys.stderr.write(f"Excel extraction failed: {str(e)}\n{traceback.format_exc()}\n")
            return ActionResponse(
                success=False,
                message=f"Excel extraction failed: {str(e)}",
                metadata={"error_type": "extraction_error"},
            )

    def mcp_create_excel_screenshot(
        self,
        file_path: str = Field(description="Path to the Excel document file"),
        sheet_name: str | None = Field(
            default=None, description="Specific sheet name to screenshot (None for first sheet)"
        ),
        max_rows: int = Field(default=50, description="Maximum number of rows to include"),
        max_cols: int = Field(default=20, description="Maximum number of columns to include"),
    ) -> ActionResponse:
        """Create a JPEG screenshot of the valid Excel area.

        This tool creates a visual representation of Excel data as a JPEG image,
        useful for further image processing or visual analysis.
        """
        try:
            # Handle FieldInfo objects
            if isinstance(file_path, FieldInfo):
                file_path = file_path.default
            if isinstance(sheet_name, FieldInfo):
                sheet_name = sheet_name.default
            if isinstance(max_rows, FieldInfo):
                max_rows = max_rows.default
            if isinstance(max_cols, FieldInfo):
                max_cols = max_cols.default

            # Validate input file
            file_path: Path = self._validate_file_path(file_path)
            print(f"Creating screenshot for Excel document: {file_path.name}")

            # Create screenshot
            screenshot_path = self._create_excel_screenshot(file_path, sheet_name)

            # Prepare metadata
            file_stats = file_path.stat()
            screenshot_stats = Path(screenshot_path).stat()

            metadata = {
                "source_file": str(file_path.absolute()),
                "source_file_size": file_stats.st_size,
                "screenshot_path": screenshot_path,
                "screenshot_size": screenshot_stats.st_size,
                "sheet_name": sheet_name,
                "max_rows_displayed": max_rows,
                "max_cols_displayed": max_cols,
                "format": "JPEG",
            }

            return ActionResponse(
                success=True,
                message=f"Excel screenshot created successfully. File saved to: {screenshot_path}",
                metadata=metadata,
            )

        except Exception as e:
            self.logger.error(f"Screenshot creation failed: {str(e)}: {traceback.format_exc()}")
            return ActionResponse(
                success=False,
                message=f"Screenshot creation failed: {str(e)}",
                metadata={"error_type": "screenshot_error"},
            )

    def mcp_list_supported_formats(self) -> ActionResponse:
        """List all supported Excel formats for extraction."""
        supported_formats = {
            "XLSX": "Excel 2007+ format files (.xlsx) - Full support including images",
            "XLS": "Excel 97-2003 format files (.xls) - Text and data only",
        }

        format_list = "\n".join(
            [f"**{format_name}**: {description}" for format_name, description in supported_formats.items()]
        )

        return ActionResponse(
            success=True,
            message=f"Supported Excel formats:\n\n{format_list}",
            metadata={"supported_formats": list(supported_formats.keys()), "total_formats": len(supported_formats)},
        )


# Example usage and entry point
if __name__ == "__main__":
    import sys
    import json
    
    # Determine if running in MCP tool mode (called without arguments)
    is_mcp_mode = len(sys.argv) == 1
    
    # Redirect print to stderr if in MCP mode
    if is_mcp_mode:
        original_print = print
        print = lambda *args, **kwargs: original_print(*args, file=sys.stderr, **kwargs)
    
    load_dotenv()

    # Default arguments for testing
    args = ActionArguments(
        name="excel_extraction_service",
        transport="stdio",
        workspace=os.getenv("MASARENA_WORKSPACE", "~"),
    )

    # Initialize and run the Excel extraction service
    try:
        service = XLSXExtractionCollection(args)
        
        # Special handling for direct stdin/stdout communication
        if is_mcp_mode:
            print("Excel extraction service ready for MCP input")
            input_line = sys.stdin.readline().strip()
            
            try:
                input_data = json.loads(input_line)
                
                # Extract tool name and arguments
                function_name = input_data.get("function_name", "extract_xlsx_content")
                arguments = input_data.get("arguments", {})
                
                # Call the appropriate tool - ensure create_screenshot is False by default
                if function_name == "extract_xlsx_content":
                    # Force create_screenshot to False for safety in server environments
                    if "create_screenshot" not in arguments:
                        arguments["create_screenshot"] = False
                        
                    result = service.mcp_extract_excel_content(
                        file_path=arguments.get("file_path", ""),
                        output_format=arguments.get("output_format", "markdown"),
                        extract_images=arguments.get("extract_images", True),
                        create_screenshot=arguments.get("create_screenshot", False),
                        sheet_names=arguments.get("sheet_names", None),
                        include_empty_cells=arguments.get("include_empty_cells", False),
                    )
                    
                    # Write result to stdout
                    sys.stdout.write(json.dumps(result.model_dump()) + "\n")
                    sys.stdout.flush()
                
                elif function_name == "extract_cell_colors":
                    result = service.mcp_extract_cell_colors(
                        file_path=arguments.get("file_path", ""),
                        sheet_name=arguments.get("sheet_name", None)
                    )
                    sys.stdout.write(json.dumps(result.model_dump()) + "\n")
                    sys.stdout.flush()
                    
                elif function_name == "list_supported_formats":
                    result = service.mcp_list_supported_formats()
                    sys.stdout.write(json.dumps(result.model_dump()) + "\n")
                    sys.stdout.flush()
                    
                else:
                    error = {"success": False, "message": f"Unknown tool: {function_name}"}
                    sys.stdout.write(json.dumps(error) + "\n")
                    sys.stdout.flush()
                    
            except json.JSONDecodeError as e:
                error = {"success": False, "message": f"Invalid JSON input: {str(e)}"}
                sys.stderr.write(f"Invalid JSON: {str(e)}\n")
                sys.stdout.write(json.dumps(error) + "\n")
                sys.stdout.flush()
                
            except Exception as e:
                error = {"success": False, "message": f"Error: {str(e)}"}
                sys.stderr.write(f"Exception: {traceback.format_exc()}\n")
                sys.stdout.write(json.dumps(error) + "\n")
                sys.stdout.flush()
        else:
            # Normal ActionCollection mode
            service.run()
    except Exception as e:
        sys.stderr.write(f"An error occurred: {e}\n{traceback.format_exc()}\n")
        sys.exit(1)
